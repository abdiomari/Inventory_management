import os
import logging
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.http import response, HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from rest_framework.renderers import JSONRenderer

from djangoProject2 import settings
from inventory_app.app_forms import ProductForm, CustomerForm, OrderForm, TransactionForm, \
    OrderDetailForm, SupplierForm, LoginForm, SignUpForm
from inventory_app.models import Product, Customer, OrderDetail, Transaction, Supplier, Order
from inventory_app.serializer import SerializeProducts, SerializeOrders

logger = logging.getLogger(__name__)


# Create your views here
@login_required
def home(request):
    products = Product.objects.all()
    paginator = Paginator(products, 5)
    page_number = request.GET.get("page")
    data = paginator.get_page(page_number)

    orders = OrderDetail.objects.all()
    customers = Customer.objects.all()
    transactions = Transaction.objects.all()

    suppliers = Supplier.objects.all()
    paginator_suppliers = Paginator(suppliers, 3)
    # page_number_suppliers = request.GET.get("page")
    suppliers = paginator_suppliers.get_page(page_number)

    total_purchases = Transaction.objects.filter(transaction_type='purchases').aggregate(
        total_purchase=Sum('amount'))['total_purchase'] or 0

    total_sales = Transaction.objects.filter(transaction_type='sales').aggregate(total_sales=Sum(
        'amount'))[
                      'total_sales'] or 0

    total_profit = total_sales - total_purchases
    total_revenue = total_sales + total_purchases

    context = {
        'products': data,
        'orders': orders,
        'customers': customers,
        'transactions': transactions,
        'suppliers': suppliers,
        "total_purchase": total_purchases,
        "total_sales": total_sales,
        "total_profit": total_profit,
        "total_revenue": total_revenue,

    }

    return render(request, "index.html", context)


@login_required
def add_customers(request):
    if request.method == "POST":
        form = CustomerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.info(request, "Added new Product")
            return redirect("home")
    else:
        form = CustomerForm()
    return render(request, "add_customer.html", {"form": form})


@login_required
def add_product(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.info(request, "Added new Product")
            return redirect("home")
    else:
        form = ProductForm()
    return render(request, "add_product.html", {"form": form})


@login_required
def view_customers(request):
    customers = Customer.objects.all()
    paginator = Paginator(customers, 20)
    page_number = request.GET.get("page")
    data = paginator.get_page(page_number)
    return render(request, "view_customers.html", {"customers": customers})


@login_required
def view_products(request):
    products = Product.objects.all()
    paginator = Paginator(products, 20)
    page_number = request.GET.get("page")
    data = paginator.get_page(page_number)
    return render(request, "view_products.html", {"products": data})


@login_required
def add_order(request):
    if request.method == "POST":
        form = OrderForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.info(request, "Added new Product")
            return redirect("home")
    else:
        form = OrderForm()
    return render(request, "add_order.html", {"form": form})


@login_required
def view_orders(request):
    orders = OrderDetail.objects.all()
    paginator = Paginator(orders, 5)
    page_number = request.GET.get("page")
    data = paginator.get_page(page_number)
    return render(request, "index.html", {"products": data})


@login_required
def add_transactions(request):
    if request.method == "POST":
        form = TransactionForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.info(request, "Added new Product")
            return redirect("home")
    else:
        form = TransactionForm()
    return render(request, "add_transaction.html", {"form": form})


@login_required
def view_transactions(request):
    transactions = Transaction.objects.all()
    paginator = Paginator(transactions, 20)
    page_number = request.GET.get("page")
    transactions = paginator.get_page(page_number)
    return render(request, "view_transactions.html", {"transactions": transactions})


@login_required
def add_order_details(request):
    if request.method == "POST":
        form = OrderDetailForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.info(request, "Added new Product")
            return redirect("home")
    else:
        form = OrderDetailForm()
    return render(request, "add_order_details.html", {"form": form})


@login_required
def view_order(request, prod_id):
    product = Product.objects.get(pk=prod_id)
    return render(request, "view_order_details.html", {"product": product})


# def product_delete(request, prod_id):
#     product = get_object_or_404(Product, pk=prod_id)
#     product.delete()
#     messages.warning(request, "This employee was deleted permanently")
#     return redirect("all")
#

def signin(request):
    if request.method == "GET":
        form = LoginForm()
        return render(request, "signin.html", {"form": form})
    elif request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                return redirect('home')
        messages.error(request, "Wrong username or password")
        return render(request, "signin.html", {"form": form})


# python man

@login_required
def signout(request):
    logout(request)
    return redirect('signin')


@login_required
def product_details(request, prod_id):
    product = Product.objects.get(pk=prod_id)
    return render(request, "view_product.html", {"product": product})


@login_required
def delete_product(request, prod_id):
    product = get_object_or_404(Product, pk=prod_id)
    product.delete()
    messages.warning(request, "This product was deleted permanently")
    return redirect("home")


@login_required
def update_products(request, prod_id):
    product = get_object_or_404(Product, pk=prod_id)
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, "Product updated successfully")
            return redirect('product_details', prod_id)
    else:
        form = ProductForm(instance=product)
    return render(request, "update_product.html", {"form": form})


@login_required
def add_supplier(request):
    if request.method == "POST":
        form = SupplierForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.info(request, "Added new Supplier")
            return redirect("home")
    else:
        form = SupplierForm()
    return render(request, "add_supplier.html", {"form": form})


@login_required
def view_suppliers(request):
    suppliers = Supplier.objects.all()
    return render(request, "view_suppliers.html", {"suppliers": suppliers})


@login_required
def view_order_details(request):
    order_details = OrderDetail.objects.all()
    return render(request, "view_order_details.html", {"suppliers": order_details})


def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('signin')
    else:
        form = SignUpForm()
    return render(request, 'signup.html', {"form": form})


def generate_excel(request):
    # Create a new Excel Workbook
    workbook = Workbook()
    excel_file = workbook.active

    headers = ['Name', 'ProductName', 'TotalAmount', 'Subtotal', 'Address']
    excel_file.append(headers)  # Add headers to the Excel file

    # Fetch Data from several Tables
    customer_table = Customer.objects.all()
    logger.info(f"Customer Data: {customer_table}")
    for row, data in enumerate(customer_table, start=1):
        excel_file.cell(row=row, column=1, value=data.name)
        excel_file.cell(row=row, column=6, value=data.address)

    order_table = Order.objects.all()
    for row, data in enumerate(order_table, start=1):
        # excel_file.cell(row=row, column=4, value=data.order_date)
        excel_file.cell(row=row, column=3, value=data.total_amount)

    order_table = Product.objects.all()
    for row, data in enumerate(order_table, start=1):
        excel_file.cell(row=row, column=2, value=data.productName)

    customer_table = OrderDetail.objects.all()
    for row, data in enumerate(customer_table, start=1):
        excel_file.cell(row=row, column=5, value=data.subtotal)
        # excel_file.cell(row=row, column=7, value=data.created_at)

    # set response headers
    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument'
    #                                      '.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename=customers_report.xlsx'

    file_path = os.path.join(settings.MEDIA_ROOT, 'customer_report.xlsx')
    # Create directories if they don't exist
    # os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # create new excel
    # workbook = Workbook()
    # excel_file = workbook.active
    workbook.save(file_path)
    # return render(request, 'view_excel.html', {"file_path": file_path})
    return JsonResponse({'message': 'Excel file generated successfully'})


def display_excel(request):
    file_path = os.path.join(settings.MEDIA_ROOT, 'customer_report.xlsx')

    # Load the Excel file
    workbook = load_workbook(file_path)
    excel_file = workbook.active

    # Read data from the Excel file
    data = []
    for row in excel_file.iter_rows(values_only=True):
        data.append(row)

    headers = data[0]  # Extract headers
    rows = data[1:]  # Extract data rows

    return render(request, 'view_excel.html', {'headers': headers, 'rows': rows})
